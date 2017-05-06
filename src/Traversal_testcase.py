# /usr/bin/env python
# coding=utf-8
import openpyxl
from framework.src.interfaceclass import  *
from framework.commons.log import log

class run_case():
    def runcase(self,test_case,save_case):
        logg = log()
        # try:
        #     wp = openpyxl.load_workbook(test_case)
        # except Exception as e:
        #     # print("打开用例失败！")
        #     logg.error("打开测试用例出错,测试用例路径：",test_case)
        wp = openpyxl.load_workbook(test_case)
        sheet = wp.get_sheet_by_name('TestCase')
        corr_dict={}

        for i in range(2,sheet.max_row+1):
            if sheet.cell(row=i,column=10).value.replace("\n","").replace("\r","") == "Yes":
                #如果用例中该字段为Yes则不执行跳出
                continue
            params = sheet.cell(row=i,column=7).value.replace("\n","").replace("\r","")
            # print("params==",params)
            url1 = sheet.cell(row=i,column=3).value.replace("\n","").replace("\r","")
            url2 = sheet.cell(row=i,column=4).value.replace("\n","").replace("\r","")
            url = url1 + url2
            requ_method = sheet.cell(row=i,column=5).value.replace("\n","").replace("\r","")
            types = sheet.cell(row=i,column=6).value.replace("\n","").replace("\r","")
            chinkoption = sheet.cell(row=i,column=8).value.replace("\n","").replace("\r","")
            headers={}
#替换请求参数
            for keyword in corr_dict:
                # print("keyword=",keyword)   #打印keyword= ${date}
                if params.find(keyword)>0:
                    # print("params=",params)  #打印params= {"key":"e711bc6362b3179f5a28de7fd3ee4ace","date":"${date}"}
                    # print(corr_dict[keyword])
                    params=params.replace(keyword,str(corr_dict[keyword]))

            params = eval(params)
            q = InterfaceTest()
            # q.testsearch(url,params,headers,requ_method,chinkoption,types,sheet,i)
            res = q.testsearch(url,params,headers,requ_method,chinkoption,types,sheet,i,logg)

            if sheet.cell(row=i,column=9).value != None:
                corr = sheet.cell(row=i,column=9).value.replace('\n','').replace('\r','')
                corr = corr.replace('\n','').replace('\r','').split(";")
                for j in range(len(corr)):
                    param = corr[j].split("=")
                    for key in param[1][1:-1].split("]["):
                        temp = res[key]
                        res = temp
                        # print(res)
                corr_dict[param[0]]=res
                # print(corr_dict)

#关联参数的拆分

        wp.save(save_case)

# test_case = "D:/pycharm workspace/framework/testcase/laohuanglitestcase5.xlsx"
# save_case = "D:/pycharm workspace/framework/report/guo.xlsx"
# run = run_case()
# run.runcase(test_case,save_case)


'''
1关联的格式
2两个关联的参数用分号；隔开
3关联参数的拆分
多次for循环进行拆分
一次是根据；进行拆分
二次是根据=进行拆分
三次是根据][进行拆分
得到key，根据key去取value值
定义一个空字典把对应的值放进去
4关联参数去替换请求中的参数
for循环去遍历key，如果找到就替换为key对应的value
'''